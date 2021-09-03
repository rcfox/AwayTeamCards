from io import BytesIO
from collections import namedtuple
import json
import sys
import time

from PIL import Image, ImageDraw, ImageFont
import cairosvg
import openpyxl

Element = namedtuple('Element', ['name', 'image', 'deck_count'])
Obstacle = namedtuple(
    'Obstacle', ['element1', 'element2', 'name', 'flavour', 'deck_count'])
Reward = namedtuple('Reward', ['name', 'text', 'deck_count', 'elements'])
Role = namedtuple('Role', ['name', 'power'])
MacGuffin = namedtuple('MacGuffin', ['name', 'power_rating', 'text'])

FONT_PATH = '/usr/share/fonts/truetype/ubuntu/Ubuntu-R.ttf'


def text_wrap(text, font, max_width):
    """Wrap text base on specified width.
        This is to enable text of width more than the image width to be display
        nicely.
        @params:
            text: str
                text to wrap
            font: obj
                font of the text
            max_width: int
                width to split the text with
        @return
            lines: list[str]
                list of sub-strings
        """
    lines = []

    # If the text width is smaller than the image width, then no need to split
    # just add it to the line list and return
    if font.getsize(text)[0] <= max_width:
        lines.append(text)
    else:
        #split the line by spaces to get words
        words = text.split(' ')
        i = 0
        # append every word to a line while its width is shorter than the image width
        while i < len(words):
            line = ''
            while i < len(words) and font.getsize(line +
                                                  words[i])[0] <= max_width:
                line = line + words[i] + " "
                i += 1
            if not line:
                line = words[i]
                i += 1
            lines.append(line)
    return '\n'.join(lines)


def parse_spreadsheet(filename):
    wb = openpyxl.load_workbook(filename, read_only=True)
    elements = []
    obstacles = []
    rewards = []
    roles = []
    macguffins = []

    try:
        for row in wb['Elements'].iter_rows(min_row=2, max_col=3):
            elements.append(Element(*[cell.value for cell in row]))

        for row in wb['Obstacles'].iter_rows(min_row=2, max_col=5):
            obstacles.append(Obstacle(*[cell.value for cell in row]))

        for row in wb['Rewards'].iter_rows(min_row=2):
            name, text, deck_count, *reward_elements = [
                cell.value for cell in row
            ]
            reward_elements = [el for el in reward_elements if el is not None]
            if text == 'Total':
                continue
            rewards.append(Reward(name, text, int(deck_count),
                                  reward_elements))

        for row in wb['SpeciesRolesTrait'].iter_rows(min_row=2,
                                                     min_col=6,
                                                     max_col=7):
            if row[0].value:
                roles.append(Role(*[cell.value for cell in row]))

        for row in wb['MacGuffins'].iter_rows(min_row=2, max_col=3):
            macguffins.append(MacGuffin(*[cell.value for cell in row]))

    finally:
        wb.close()

    return elements, obstacles, rewards, roles, macguffins


def svg2image(svg, size):
    width, height = size
    out = BytesIO()
    cairosvg.svg2png(url=f'icons/{svg}',
                     write_to=out,
                     parent_width=width,
                     parent_height=height)
    return Image.open(out)


def card_template(title, card_type, text):
    width = 407
    height = 585

    num_text_lines = 8

    inset = 24
    box_separation = 8

    rect_radius = 8
    rect_outline_width = 3

    title_font = ImageFont.truetype(FONT_PATH, size=32)
    title_padding = 4

    type_font = ImageFont.truetype(FONT_PATH, size=16)
    type_padding = 2

    text_font = ImageFont.truetype(FONT_PATH, size=24)
    text_padding = 8

    img = Image.new('RGBA', (width, height), color=(255, 255, 255, 255))
    draw = ImageDraw.Draw(img)

    # ---

    # Add 'Q' to every string to try to get a uniform text height.
    _, title_height = title_font.getsize(title + 'Q')
    title_box_height = title_height + title_padding * 2
    draw.rounded_rectangle(
        ((inset, inset), (width - inset, inset + title_box_height)),
        radius=rect_radius,
        outline=(0, 0, 0),
        width=rect_outline_width)

    title_x = inset + rect_radius + title_padding
    title_y = inset + title_padding
    draw.text((title_x, title_y),
              title,
              font=title_font,
              fill=(0, 0, 0),
              anchor='la')

    # ---

    type_width, type_height = type_font.getsize(card_type + 'Q')
    type_x = width - inset
    type_y = height - inset - type_height // 2 - type_padding
    draw.text((type_x, type_y),
              card_type,
              font=type_font,
              fill=(0, 0, 0),
              anchor='rm')

    type_box_x = width // 2
    type_box_y = height - inset
    type_box_width = type_width + type_padding * 2
    type_box_height = type_height + type_padding * 2
    # draw.rounded_rectangle(
    #     ((type_box_x - type_box_width // 2, type_box_y - type_box_height),
    #      (type_box_x + type_box_width // 2, type_box_y)),
    #     radius=rect_radius,
    #     outline=(0, 0, 0),
    #     width=rect_outline_width)

    # ---

    text_box_y = type_box_y - type_box_height
    text_box_height = 0
    if text:
        text_width, text_height = text_font.getsize(text + 'Q')
        text_box_y = type_box_y - type_box_height - box_separation
        text_box_height = num_text_lines * text_height + text_padding * 2
        draw.rounded_rectangle(((inset, text_box_y - text_box_height),
                                (width - inset, text_box_y)),
                               radius=rect_radius,
                               outline=(0, 0, 0),
                               width=rect_outline_width)

        draw.multiline_text((width // 2, text_box_y - text_box_height // 2),
                            text_wrap(text, text_font,
                                      width - inset * 2 - text_padding * 2),
                            font=text_font,
                            fill=(0, 0, 0),
                            anchor='mm',
                            align='center')

    # --

    img_box_y = inset + title_height + title_padding * 2 + box_separation
    img_box_height = text_box_y - img_box_y - text_box_height - box_separation
    draw.rounded_rectangle(
        ((inset, img_box_y), (width - inset, img_box_y + img_box_height)),
        radius=rect_radius,
        outline=(0, 0, 0),
        width=rect_outline_width)

    return img, (inset, img_box_y), (width - inset, img_box_y + img_box_height)


def card_template2(title, card_type, text):
    width = 407
    height = 585

    num_text_lines = 16

    inset = 24
    box_separation = 8

    rect_radius = 8
    rect_outline_width = 3

    title_font = ImageFont.truetype(FONT_PATH, size=32)
    title_padding = 4

    type_font = ImageFont.truetype(FONT_PATH, size=16)
    type_padding = 2

    text_font = ImageFont.truetype(FONT_PATH, size=24)
    text_padding = 8

    img = Image.new('RGBA', (width, height), color=(255, 255, 255, 255))
    draw = ImageDraw.Draw(img)

    # ---

    # Add 'Q' to every string to try to get a uniform text height.
    _, title_height = title_font.getsize(title + 'Q')
    title_box_height = title_height + title_padding * 2
    draw.rounded_rectangle(
        ((inset, inset), (width - inset, inset + title_box_height)),
        radius=rect_radius,
        outline=(0, 0, 0),
        width=rect_outline_width)

    title_x = inset + rect_radius + title_padding
    title_y = inset + title_padding
    draw.text((title_x, title_y),
              title,
              font=title_font,
              fill=(0, 0, 0),
              anchor='la')

    # ---

    type_width, type_height = type_font.getsize(card_type + 'Q')
    type_x = width - inset
    type_y = height - inset - type_height // 2 - type_padding
    draw.text((type_x, type_y),
              card_type,
              font=type_font,
              fill=(0, 0, 0),
              anchor='rm')

    type_box_x = width // 2
    type_box_y = height - inset
    type_box_width = type_width + type_padding * 2
    type_box_height = type_height + type_padding * 2
    # draw.rounded_rectangle(
    #     ((type_box_x - type_box_width // 2, type_box_y - type_box_height),
    #      (type_box_x + type_box_width // 2, type_box_y)),
    #     radius=rect_radius,
    #     outline=(0, 0, 0),
    #     width=rect_outline_width)

    # ---

    text_box_y = type_box_y - type_box_height
    text_box_height = 0
    if text:
        text_width, text_height = text_font.getsize(text + 'Q')
        text_box_y = type_box_y - type_box_height - box_separation
        text_box_height = num_text_lines * text_height + text_padding * 2
        draw.rounded_rectangle(((inset, text_box_y - text_box_height),
                                (width - inset, text_box_y)),
                               radius=rect_radius,
                               outline=(0, 0, 0),
                               width=rect_outline_width)

        draw.multiline_text((width // 2, text_box_y - text_box_height // 2),
                            text_wrap(text, text_font,
                                      width - inset * 2 - text_padding * 2),
                            font=text_font,
                            fill=(0, 0, 0),
                            anchor='mm',
                            align='center')

    # --

    img_box_y = inset + title_height + title_padding * 2 + box_separation
    img_box_height = text_box_y - img_box_y - text_box_height - box_separation
    draw.rounded_rectangle(
        ((inset, img_box_y), (width - inset, img_box_y + img_box_height)),
        radius=rect_radius,
        outline=(0, 0, 0),
        width=rect_outline_width)

    return img, (inset, img_box_y), (width - inset, img_box_y + img_box_height)


def icons_vertical(img, xy1, xy2, icons):
    x1, y1 = xy1
    x2, y2 = xy2

    width = x2 - x1
    height = y2 - y1

    icon_size = min(width, height) // (len(icons) + 1)
    icon_x = x1 + width // 2 - icon_size // 2

    y_step = (height - icon_size) // len(icons)

    icon_imgs = [svg2image(icon, (icon_size, icon_size)) for icon in icons]

    icon_y = y1 + height // (len(icons) + 1) - icon_size // 2
    for icon_img in icon_imgs:
        img.paste(icon_img, (icon_x, icon_y), icon_img)
        icon_y += y_step

    return img


def icons_horizontal(img, xy1, xy2, icons):
    x1, y1 = xy1
    x2, y2 = xy2

    width = x2 - x1
    height = y2 - y1

    icon_size = min(width, height) // max(len(icons), 2)
    icon_y = y1 + height // 2 - icon_size // 2

    x_step = (width - icon_size) // len(icons)

    icon_imgs = [svg2image(icon, (icon_size, icon_size)) for icon in icons]

    icon_x = x1 + width // (len(icons) + 1) - icon_size // 2
    for icon_img in icon_imgs:
        img.paste(icon_img, (icon_x, icon_y), icon_img)
        icon_x += x_step

    return img


def draw_element_card(element, icon):
    img, xy1, xy2 = card_template(element, 'ELEMENT', '')
    return icons_vertical(img, xy1, xy2, [icon])


def draw_obstacle_card(obstacle, elements):
    img, xy1, xy2 = card_template(obstacle.name, 'OBSTACLE', obstacle.flavour)
    icons = [elements[obstacle.element1], elements[obstacle.element2]]
    return icons_horizontal(img, xy1, xy2, icons)


def draw_reward_card(reward, elements):
    name = reward.name
    if not name:
        name = '/'.join(reward.elements)

    img, xy1, xy2 = card_template(name, 'REWARD', reward.text)

    icons = [elements[el] for el in reward.elements]

    return icons_vertical(img, xy1, xy2, icons)


def draw_role_card(role):
    img, xy1, xy2 = card_template(role.name, 'ROLE', role.power)
    return icons_vertical(img, xy1, xy2, ['role.svg'])


def draw_macguffin_card(macguffin):
    img, xy1, xy2 = card_template2(
        f'({macguffin.power_rating}) {macguffin.name}', 'MACGUFFIN',
        macguffin.text)
    return img


def create_image_sheets(name, images, hidden):
    if len(images) > 69:
        yield from create_image_sheets(name + '_', images[69:], hidden)

    images = images[:69]

    sheet_width = 10
    if len(images) <= 10:
        sheet_width = len(images) // 2 + 1
    sheet_height = len(images) // sheet_width + 1

    card_width, card_height = hidden.size

    sheet = Image.new('RGB',
                      (sheet_width * card_width, sheet_height * card_height))
    for idx, image in enumerate(images):
        x = (idx % sheet_width) * card_width
        y = (idx // sheet_width) * card_height
        sheet.paste(image, (x, y), image)

    sheet.paste(hidden, ((sheet_width - 1) * card_width,
                         (sheet_height - 1) * card_height), hidden)
    sheet.save(f'generated/{name}.png')

    yield name, min(len(images), 69), sheet_width, sheet_height


def hidden_card():
    img, xy1, xy2 = card_template('???', 'HIDDEN', '')
    return icons_vertical(img, xy1, xy2, ['hidden.svg'])


def base_transform():
    return {
        'posX': 0,
        'posY': 0,
        'posZ': 0,
        'rotX': 0,
        'rotY': 180,
        'rotZ': 180,
        'scaleX': 1,
        'scaleY': 1,
        'scaleZ': 1
    }


def collection_json(*items):
    return {
        'ObjectStates': items,
    }


def deck_json(name, description):
    return {
        'Name': 'DeckCustom',
        'Transform': base_transform(),
        'DeckIDs': [],
        'Nickname': name,
        'Description': description,
        'Hands': False,  # Don't allow decks to go to player hands
        'SidewaysCard': False,  # Cards aren't meant to be held sideways
        'CustomDeck': {},
        'ContainedObjects': [],
    }


def subdeck_json(deck_img, backface_url, width, height):
    base_url = 'https://raw.githubusercontent.com/rcfox/AwayTeamCards/master/generated/{}.png?{}'
    return {
        'FaceURL': base_url.format(deck_img, int(time.time())),
        'BackURL': backface_url,
        'NumWidth': width,
        'NumHeight': height,
        'BackIsHidden': False,  # Don't use back face as the hidden card
        'UniqueBack': False,  # BackURL isn't a card sheet
        'Type': 0,  # Rectangle
    }


def card_json(name, description, deck_count):
    return {
        'Name': 'Card',
        'Transform': base_transform(),
        'Nickname': name,
        'Description': description,
        'CardID': 0,
        'Hands': True,  # Allow in player hands
        'SidewaysCard': False,  # Not sideways
        '__count__': deck_count,
    }


def define_deck(deck, deck_name, backface_url, card_data, card_imgs,
                hidden_card):
    for idx, sheet_data in enumerate(create_image_sheets(
            deck_name, card_imgs, hidden_card),
                                     start=10):
        name, count, width, height = sheet_data
        cards = card_data[:count]
        card_data = card_data[count:]
        for card_id, card in enumerate(cards):
            deck_count = int(card.pop('__count__', 1))

            card['CardID'] = int(f'{idx}{card_id:02d}')
            for _ in range(deck_count):
                deck['DeckIDs'].append(card['CardID'])
                deck['ContainedObjects'].append(card)

        subdeck = subdeck_json(name, backface_url, width, height)
        deck['CustomDeck'][str(idx)] = subdeck

    #with open(f'generated/{deck_name}.json', 'w') as f:
    #    json.dump(collection_json(deck), f, indent=2)
    return deck


def define_element_deck(elements, hidden_card):
    backface_url = 'https://i.imgur.com/A7Wa5GN.jpeg'

    card_imgs = []
    card_data = []
    for element in elements:
        if element.name is None:
            continue

        card_data.append(card_json(element.name, '', element.deck_count))
        card_imgs.append(draw_element_card(element.name, element.image))

    deck = deck_json('Elements', 'Use these to overcome obstacles.')

    return define_deck(deck, 'element', backface_url, card_data, card_imgs,
                       hidden_card)


def define_obstacle_deck(obstacles, element_icons, hidden_card):
    backface_url = 'https://i.imgur.com/LS2eGhT.jpg'

    card_imgs = []
    card_data = []
    for obstacle in obstacles:
        if obstacle.name is None:
            continue

        card_data.append(
            card_json(obstacle.name, obstacle.flavour, obstacle.deck_count))
        card_imgs.append(draw_obstacle_card(obstacle, element_icons))

    deck = deck_json('Obstacles', 'Overcome these to win rewards.')

    return define_deck(deck, 'obstacle', backface_url, card_data, card_imgs,
                       hidden_card)


def define_reward_deck(rewards, element_icons, hidden_card):
    backface_url = 'https://i.imgur.com/qtH7MDU.jpg'

    card_imgs = []
    card_data = []
    for reward in rewards:
        if reward.text == 'Total':
            continue

        card_data.append(card_json(reward.name, reward.text,
                                   reward.deck_count))
        card_imgs.append(draw_reward_card(reward, element_icons))

    deck = deck_json(
        'Rewards',
        'Bonuses to help overcome obstacles or give special abilities.')

    return define_deck(deck, 'reward', backface_url, card_data, card_imgs,
                       hidden_card)


def define_role_deck(roles, hidden_card):
    backface_url = 'https://hdwallpaperim.com/wp-content/uploads/2017/08/22/453064-space-space_art-stars-planet-nebula-galaxy.jpg'

    card_imgs = []
    card_data = []
    for role in roles:
        if role.name is None:
            continue

        card_data.append(card_json(role.name, role.power, 1))
        card_imgs.append(draw_role_card(role))

    deck = deck_json(
        'Roles', 'Player-specific special abilities to help you in the game.')

    return define_deck(deck, 'role', backface_url, card_data, card_imgs,
                       hidden_card)


def define_macguffin_deck(macguffins, hidden_card):
    backface_url = 'https://i.imgur.com/4FsVPIC.jpg'

    card_imgs = []
    card_data = []
    for macguffin in macguffins:
        if macguffin.name is None:
            continue

        card_data.append(card_json(macguffin.name, macguffin.text, 1))
        card_imgs.append(draw_macguffin_card(macguffin))

    deck = deck_json('MacGuffins', 'Problems to solve.')

    return define_deck(deck, 'macguffin', backface_url, card_data, card_imgs,
                       hidden_card)


def main():
    elements, obstacles, rewards, roles, macguffins = parse_spreadsheet(
        sys.argv[1])

    element_icons = {element.name: element.image for element in elements}

    hidden = hidden_card()

    collection = collection_json(
        define_element_deck(elements, hidden),
        define_obstacle_deck(obstacles, element_icons, hidden),
        define_reward_deck(rewards, element_icons, hidden),
        define_role_deck(roles, hidden),
        define_macguffin_deck(macguffins, hidden))

    for idx, deck in enumerate(collection['ObjectStates']):
        deck['Transform']['posX'] = idx * 2.5

    with open('all.json', 'w') as f:
        json.dump(collection, f, indent=2)


if __name__ == '__main__':
    main()
